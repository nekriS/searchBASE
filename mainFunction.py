import os
import pandas as pd
import utility as ut
import system as st
import openpyxl as oxl
import threading
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import datetime

MINIMAL_LEN = 2

class options:
    def __init__(self, 
                 bom_use_columns = "A, B, C, D, E, F", # необходимые столбцы из перечня компонентов
                 bom_skip_rows = 8, # сколько строк при анализе необходимо удалить
                 bom_pn_number = 3, # номер столбца, в котором находится парт номер
                 bom_type_number = 5, # номер столбца, в котором находится тип монтажа
                 save_bom2excel = False, # 
                 name_bom2excel = "bom.xlsx", # 
                 base_table_number = 1, # номер таблицы из базы
                 base_pn_number = 0, # номер столбца, в котором находится парт номер
                 base_drop_column = [11, 10, 9, 8, 4, 3, 2], # номера столбцов, которые необходимо выкинуть
                 base_balance_number = 2, # 
                 base_1c_number = 3, # 
                 base_comm_number = 1, # 
                 save_base2excel = False, # сохранить ли базу в эксель
                 name_base2excel = "base.xlsx", # название для базы при сохранении в эксель
                 compare_method = "normal", # тип сравнения нормальный (с порогами) или с помощью нейронной сети
                 quiet_mode = False, # включает тихий режим (отключает сообщения отладки)
                 log_object = -1):
        self.bom_use_columns = bom_use_columns 
        self.bom_skip_rows = bom_skip_rows 
        self.bom_pn_number = bom_pn_number 
        self.bom_type_number = bom_type_number
        self.save_bom2excel = save_bom2excel
        self.name_bom2excel = name_bom2excel
        self.base_table_number = base_table_number
        self.base_pn_number = base_pn_number
        self.base_drop_column = base_drop_column
        self.base_balance_number = base_balance_number
        self.base_1c_number = base_1c_number 
        self.base_comm_number = base_comm_number 
        self.save_base2excel = save_base2excel
        self.name_base2excel = name_base2excel
        self.compare_method = compare_method
        self.quiet_mode = quiet_mode
        self.log_object = log_object



def find_bom_in_base(name_bom, name_base, options):

    st.log("The process has begun.", options.log_object)
    start_time = datetime.datetime.now()
    params = ', '.join(f"{k}={repr(v)}" for k, v in vars(options).items())
    st.log("Options: " + params, options.log_object)

    
    bom_table = pd.read_excel(name_bom, usecols=options.bom_use_columns, skiprows=options.bom_skip_rows)
    header_bom = bom_table.columns.tolist()
    name_pn_bom = header_bom[options.bom_pn_number]
    type_part_bom = header_bom[options.bom_type_number]

    base_table = pd.read_html(name_base, encoding='cp1251')[options.base_table_number]
    header_base = base_table.columns.tolist()
    for i_column in options.base_drop_column:
        base_table = base_table.drop(columns=[header_base[i_column]])
    #base_table = base_table.drop(options.base_drop_column)
    header_base = base_table.columns.tolist()
    name_pn_base = header_base[options.base_pn_number]
    name_comm_base = header_base[options.base_comm_number]
    name_1c_base = header_base[options.base_1c_number]
    name_balance = header_base[options.base_balance_number]

    #print(base_table)
    #print(header_base)

    if options.save_base2excel:
        base_table.to_excel(options.name_base2excel, index=False)

    bom_table['SMT'] = ""
    bom_table['SMT Баланс'] = 0
    bom_table['SMT Жаккара'] = ""
    bom_table['SMT Левенштейна'] = ""
    bom_table['Результат'] = 0

    for index_bom, row_bom in bom_table.iterrows():
        
        str1 = str(row_bom[name_pn_bom]).replace("nan", "")
        type_part = str(row_bom[type_part_bom]).replace("nan", "")




        result = 0
        balance = 0
        name_in_base = ""

        components_Lev = []
        components_Jac = []
        components_All = []

        if (len(type_part) > MINIMAL_LEN) and (len(str1) > MINIMAL_LEN):
            for index_base, row_base in base_table.iterrows():

                balance = row_base[name_balance]
                str2 = str(row_base[name_pn_base])
                str2_comm = str(row_base[name_comm_base])
                str2_1c = str(row_base[name_1c_base])

                strs = [str2, str2_comm, str2_1c]
                coef_Lev, i_Lev = ut.nmax([ut.compare(str1, str2, 'Levenshtein'), ut.compare(str1, str2_comm, 'Levenshtein'), ut.compare(str1, str2_1c, 'Levenshtein')])
                coef_Jac, i_Jac = ut.nmax([ut.compare(str1, str2, 'Jacquard'), ut.compare(str1, str2_comm, 'Jacquard'), ut.compare(str1, str2_1c, 'Jacquard')])
                coef_equ, i_equ =  ut.nmax([ut.compare(str1, str2), ut.compare(str1, str2_comm), ut.compare(str1, str2_1c)])

                if coef_equ == 1:
                    if ut.compare(str1, str2, 'Levenshtein') > 0.9 and ut.compare(str1, str2, 'Jacquard') > 0.6:
                        result = 2
                        name_in_base = str2
                        break
                    elif ut.compare(str1, str2, 'Jacquard') > 0.3:
                        result = 1
                        name_in_base = str2
                        break
                if (coef_Lev > 0.5) and (coef_Jac > 0.5): 
                    components_All.append([coef_Lev, coef_Jac, str2, strs[i_Lev]])
                else:
                    if (coef_Lev > 0.5):
                        if ut.isRLC(str1)[0]:
                            if ut.isRLC(str1)[1] in str2:
                                components_Lev.append([coef_Lev, str2, strs[i_Lev]])
                        else:
                            components_Lev.append([coef_Lev, str2, strs[i_Lev]])
                    if (coef_Jac > 0.5):
                        components_Jac.append([coef_Jac, str2, strs[i_Jac]])

            components_Lev = sorted(components_Lev, key=lambda x: x[0], reverse=True)
            components_Jac = sorted(components_Jac, key=lambda x: x[0], reverse=True)
            components_All = sorted(components_All, key=lambda x: x[0], reverse=True)

            if result == 0:
                if options.compare_method == "normal":
                    if (len(components_All) > 0):
                        if components_All[0][0] > 0.9 and components_All[0][1] > 0.6:
                            result = 1
                            name_in_base = str(components_All[0][2])
                        else:
                            if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                                if components_Lev[0][0] > 0.95:
                                    result = 1
                                    name_in_base = str(components_Lev[0][1])
                                else:
                                    result = -1
                            else:
                                result = -2
                    else:
                        if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                            if components_Lev[0][0] > 0.95:
                                result = 1
                                name_in_base = str(components_Lev[0][1])
                            else:
                                result = -1
                        else:
                            result = -2

            if result < 0:
                bom_table.at[index_bom, 'SMT Левенштейна'] = str(components_Lev)
                bom_table.at[index_bom, 'SMT Жаккара'] = str(components_Jac)

            bom_table.at[index_bom, 'SMT'] = name_in_base
            bom_table.at[index_bom, 'Результат'] = result
            bom_table.at[index_bom, 'SMT Баланс'] = balance

        if not(options.quiet_mode):
            if str1 != "" and str1 != " ":
                st.log(f"{str1} res={str(result)} bal={str(balance)}.", options.log_object)

    end_time = datetime.datetime.now()
    deltatime_str = st.get_time_difference(start_time, end_time)
    st.log(f"The BOM file with results has been generated! Lead time: {deltatime_str}", options.log_object)
    if options.save_bom2excel:
        bom_table.to_excel(options.name_bom2excel, index=False)
        st.log(f"The BOM file with the results was saved as {options.name_bom2excel}.", options.log_object)


    return bom_table
        


def draw_file(name_bom, bom_table, outputname="output.xlsx", open_file=False):

    workbook = oxl.load_workbook(name_bom)
    sheet_names = workbook.sheetnames
    main_sheet = workbook.worksheets[0]
    main_sheet.title = "Перечень элементов"

    color2_fill = PatternFill(start_color="97e253", end_color="97e253", fill_type="solid")
    color1_fill = PatternFill(start_color="ccf1aa", end_color="ccf1aa", fill_type="solid") 
    color0_fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    color_1_fill = PatternFill(start_color="ffd889", end_color="ffd889", fill_type="solid")
    color_2_fill = PatternFill(start_color="f4a190", end_color="f4a190", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),  # Левая граница
        right=Side(style="thin"),  # Правая граница
        top=Side(style="thin"),  # Верхняя граница
        bottom=Side(style="thin")  # Нижняя граница
    )

    ut.move_column(main_sheet, 12, 1)
    main_sheet.cell(row=9, column=13).alignment = alignment_center

    main_sheet.cell(row=9, column=11).value = "SMT"
    main_sheet.cell(row=9, column=11).border = thin_border

    main_sheet.cell(row=9, column=12).value = "SMT Баланс"
    main_sheet.cell(row=9, column=12).border = thin_border

    table_column = 9

    main_sheet.cell(row=3, column=table_column).fill = color2_fill
    main_sheet.cell(row=3, column=table_column+1).value = "Наименование есть в SMT"
    main_sheet.cell(row=3, column=table_column+1).alignment = alignment_left
    main_sheet.cell(row=4, column=table_column).fill = color1_fill
    main_sheet.cell(row=4, column=table_column+1).value = "Наименование есть в SMT, но написание отличается"
    main_sheet.cell(row=4, column=table_column+1).alignment = alignment_left
    main_sheet.cell(row=5, column=table_column).fill = color0_fill
    main_sheet.cell(row=5, column=table_column+1).value = "Не анализируется"
    main_sheet.cell(row=5, column=table_column+1).alignment = alignment_left
    main_sheet.cell(row=6, column=table_column).fill = color_1_fill
    main_sheet.cell(row=6, column=table_column+1).value = "Наименование не найдено в SMT, но есть похожие"
    main_sheet.cell(row=6, column=table_column+1).alignment = alignment_left
    main_sheet.cell(row=7, column=table_column).fill = color_2_fill
    main_sheet.cell(row=7, column=table_column+1).value = "Наименование не найдено в SMT"
    main_sheet.cell(row=7, column=table_column+1).alignment = alignment_left

    for i in range(3, 8):
        for j in range(table_column, 12):
            main_sheet.cell(row=i, column=j).border = thin_border

        main_sheet.merge_cells(f'J{i}:L{i}')

    for index_bom, row_bom in bom_table.iterrows():

        main_sheet.cell(row=index_bom + 10, column=11).value = bom_table.at[index_bom, 'SMT']
        main_sheet.cell(row=index_bom + 10, column=11).alignment = alignment_center
        main_sheet.cell(row=index_bom + 10, column=11).border = thin_border

        main_sheet.cell(row=index_bom + 10, column=12).value = bom_table.at[index_bom, 'SMT Баланс']
        main_sheet.cell(row=index_bom + 10, column=12).alignment = alignment_center
        main_sheet.cell(row=index_bom + 10, column=12).border = thin_border

        main_sheet.cell(row=index_bom + 10, column=13).alignment = alignment_center

        match bom_table.at[index_bom, 'Результат']:
            case -1:
                new_sheet = workbook.create_sheet(title=row_bom['Номенклатура'])
                
                hyperlink_cell1 = new_sheet.cell(row=1, column=1)  
                hyperlink_cell1.value = "На главную"
                hyperlink_cell1.hyperlink = f"#'{main_sheet.title}'!K{index_bom+10}" 
                hyperlink_cell1.font = Font(underline="single", color="0563C1")  
                new_sheet.cell(row=1, column=1).alignment = alignment_center

                hyperlink_cell2 = main_sheet.cell(row=index_bom + 10, column=11)  
                hyperlink_cell2.value = str(eval(row_bom['SMT Левенштейна'])[0][1]) #"Похожие"
                hyperlink_cell2.hyperlink = f"#'{row_bom['Номенклатура']}'!A1" 
                hyperlink_cell2.font = Font(underline="single", color="0563C1") 
                row = 3
                new_sheet.cell(row=row, column=1).value = "Референс:"
                new_sheet.cell(row=row, column=1).font = Font(bold=True)
                new_sheet.cell(row=row, column=1).alignment = alignment_center
                new_sheet.cell(row=row, column=1).border = thin_border
                new_sheet.cell(row=row, column=2).value = str(row_bom['Номенклатура'])
                new_sheet.cell(row=row, column=2).alignment = alignment_center
                new_sheet.cell(row=row, column=2).border = thin_border
                new_sheet.row_dimensions[row].height = 25.5
                row = 5
                new_sheet.cell(row=row, column=1).value = "Коэффициент Левенштейна"
                new_sheet.cell(row=row, column=2).value = "SMT"
                new_sheet.cell(row=row, column=1).font = Font(bold=True)
                new_sheet.cell(row=row, column=2).font = Font(bold=True)
                new_sheet.cell(row=row, column=1).alignment = alignment_center
                new_sheet.cell(row=row, column=1).border = thin_border
                new_sheet.cell(row=row, column=2).alignment = alignment_center
                new_sheet.cell(row=row, column=2).border = thin_border
                new_sheet.row_dimensions[row].height = 25.5
                
                row += 1
                for elem in eval(row_bom['SMT Левенштейна']):
                    new_sheet.cell(row=row, column=1).value = elem[0]
                    new_sheet.cell(row=row, column=2).value = elem[1]
                    
                    new_sheet.cell(row=row, column=1).alignment = alignment_center
                    new_sheet.cell(row=row, column=1).border = thin_border
                    new_sheet.cell(row=row, column=2).alignment = alignment_center
                    new_sheet.cell(row=row, column=2).border = thin_border

                    row += 1

                ut.set_column_autowidth(new_sheet, ['A', 'B'])
            case -2:
                pass

        for i in range(1, 13):
            match bom_table.at[index_bom, 'Результат']:
                case 2:
                    main_sheet.cell(row=index_bom + 10, column=i).fill = color2_fill
                case 1:
                    main_sheet.cell(row=index_bom + 10, column=i).fill = color1_fill
                case 0:
                    main_sheet.cell(row=index_bom + 10, column=i).fill = color0_fill
                case -1:
                    main_sheet.cell(row=index_bom + 10, column=i).fill = color_1_fill
                case -2:
                    main_sheet.cell(row=index_bom + 10, column=i).fill = color_2_fill



    ut.set_column_autowidth(main_sheet, ['K', 'L', 'M'])

    outputname = st.get_name_file(outputname)

    workbook.save(outputname)
    if open_file:
        try:
            os.startfile(outputname)
        except:
            pass
            #log("Не удалось открыть полученный файл!", log_object)


if __name__ == "__main__":

    name_bom = ""
    name_base = ""
    options_1 = options()

    bom_table = find_bom_in_base(name_bom, name_base, options_1)
    draw_file(name_bom, bom_table)

from PySide6.QtCore import QAbstractTableModel, Qt
from PySide6.QtGui import QColor

import os
import pandas as pd
import utility as ut
import system as st
import openpyxl as oxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import configparser
from openpyxl.utils import get_column_letter
import ast
import datetime as dt
import fnmatch
import time

MINIMAL_LEN = 2
NAME_CONFIG_FILE = "config.ini"
DEFAULT_CONFIG = {
    'LIST_DEFAULT': {
        'view_name': "По умолчанию",
        'list_use_columns': "A, B, C, D, E, F",
        'list_skip_rows': 8,
        'list_pn_number': 3,
        'check_empty_type_number': True,
        'no_montage': 'NM',
        'list_type_number': 5,
        'save_list_after_check': False,
        'name_list_after_check': "list_after_check.xlsx",
        'frame_visible': True,
        'moves': True,
        'start_custom_column': 11,
        'widths': True,
        'cap_mask': "TYPE-SIZE-DIE-VOL-VALUE TOL",
        'res_mask': "TYPE-SIZE-VALUE TOL"
    },
    'LIST_SDC': {
        'view_name': "SDC",
        'list_use_columns': "A, B, C, E, F, H, M",
        'list_skip_rows': 0,
        'list_pn_number': 4,
        'check_empty_type_number': False,
        'no_montage': 'DNP',
        'list_type_number': 5,
        'save_list_after_check': False,
        'name_list_after_check': "list_after_check.xlsx",
        'frame_visible': False,
        'moves': False,
        'start_custom_column': 16,
        'widths': False,
        'cap_mask': "SIZE-DIE-VOL-VALUE TOL",
        'res_mask': "SIZE-VALUE TOL"
    },
    'BASE_DEFAULT': {
        'view_name': "По умолчанию",
        'base_table_number': 2,
        'base_pn_number': 0,
        'base_drop_column': [11, 10, 9, 8, 4, 3, 2],
        'base_balance_number': 2,
        'base_1c_number': 3,
        'base_comm_number': 1,
        'save_base2excel': False,
        'name_base2excel': "base.xlsx",
        'search_column': 3
    },
    'GENERAL': {
        'compare_method': "normal",
        'quiet_mode': False
    }
}

def print_config(config):
    """
    Форматировано выводит содержимое объекта ConfigParser в консоль.
    
    :param config: Объект configparser.ConfigParser
    """
    for section in config.sections():
        print(f"[{section}]")
        for key, value in config[section].items():
            print(f"{key} = {value}")
        print()

class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])


        value = self._data.iloc[index.row(), index.column()]
        if role == Qt.BackgroundRole:
            try:

                #if isinstance(str(value), str) and "R-0402" in str(value):
                #    return QColor("#ffeeaa")

                if isinstance(float(value), (int, float)) and float(value) <= 0:
                    return QColor("#FFBCBC")

            except Exception:
                pass

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(self._data.columns[section])
            elif orientation == Qt.Vertical:
                return str(self._data.index[section])
        return None

class options():
    def __init__(self,
                 config = configparser.ConfigParser(),
                 check_hand = False,
                 check_nm = False,
                 log_object = -1):
        self.config = config
        self.check_hand = check_hand
        self.check_nm = check_nm
        self.log_object = log_object

        if os.path.exists(NAME_CONFIG_FILE):
            self.update()
        else:
            for section, value in DEFAULT_CONFIG.items():
                self.config[section] = value

            with open(NAME_CONFIG_FILE, "w", encoding="utf-8") as f:
                self.config.write(f)

    def update(self):
        if os.path.exists(NAME_CONFIG_FILE):
            self.config.read(NAME_CONFIG_FILE, encoding="utf-8")


def add_config_to_class(class_object, section):
    for param, value in class_object.config[section].items():
        try:
            setattr(class_object, param, int(value))
        except:
            try:
                setattr(class_object, param, ast.literal_eval(value))
            except:
                setattr(class_object, param, value)
    return class_object

def find_bom_in_base(name_bom, name_base, options, preset_base='BASE_DEFAULT', preset_list='LIST_DEFAULT'):

    options = add_config_to_class(options, preset_base)
    options = add_config_to_class(options, preset_list)
    options = add_config_to_class(options, "GENERAL")

    st.log("The process find LIST in table has begun.", options.log_object)
    start_time = dt.datetime.now()
    params = ', '.join(f"{k}={repr(v)}" for k, v in vars(options).items())
    st.log(f"Options: name_bom={name_bom}, name_base={name_base}, {params}", options.log_object)

    bom_table = pd.read_excel(name_bom, usecols=options.list_use_columns, skiprows=options.list_skip_rows)

    header_bom = bom_table.columns.tolist()
    name_pn_bom = header_bom[options.list_pn_number]
    type_part_bom = header_bom[options.list_type_number]

    base_table = pd.read_html(name_base, encoding='cp1251')[options.base_table_number]
    header_base = base_table.columns.tolist()
    for i_column in options.base_drop_column:
        base_table = base_table.drop(columns=[header_base[i_column]])

    header_base = base_table.columns.tolist()
    name_pn_base = header_base[options.base_pn_number]
    name_comm_base = header_base[options.base_comm_number]
    name_1c_base = header_base[options.base_1c_number]
    name_balance = header_base[options.base_balance_number]

    if options.save_base2excel:
        base_table.to_excel(options.name_base2excel, index=False)

    bom_table['SMT'] = ""
    bom_table['SMT Баланс'] = 0
    bom_table['SMT Жаккара'] = ""
    bom_table['SMT Левенштейна'] = ""
    bom_table['Результат'] = 0
    i = 0

    for index_bom, row_bom in bom_table.iterrows():
        
        str1 = str(row_bom[name_pn_bom]).replace("nan", "")
        type_part = str(row_bom[type_part_bom]).replace("nan", "")

        result = 0
        balance = 0
        name_in_base = ""

        components_Lev = []
        components_Jac = []
        components_All = []
        
        if type_part == "HAND" and not(options.check_hand):
            continue
        if type_part == options.no_montage and not(options.check_nm):
            continue
        if options.check_empty_type_number and len(type_part) <= 1:
            continue

        if len(str1) > MINIMAL_LEN:
            for index_base, row_base in base_table.iterrows():

                balance = row_base[name_balance]
                str2 = str(row_base[name_pn_base])
                str2_comm = str(row_base[name_comm_base])
                str2_1c = str(row_base[name_1c_base])

                match options.search_column:
                    case options.base_pn_number:
                        str_target = str2
                    case options.base_comm_number:
                        str_target = str2_comm
                    case options.base_1c_number:
                        str_target = str2_1c

                strs = [str2, str2_comm, str2_1c]
                coef_Lev, i_Lev = ut.nmax([ut.compare(str1, str2, 'Levenshtein'), ut.compare(str1, str2_comm, 'Levenshtein'), ut.compare(str1, str2_1c, 'Levenshtein')])
                coef_Jac, i_Jac = ut.nmax([ut.compare(str1, str2, 'Jacquard'), ut.compare(str1, str2_comm, 'Jacquard'), ut.compare(str1, str2_1c, 'Jacquard')])
                coef_equ, i_equ =  ut.nmax([ut.compare(str1, str2), ut.compare(str1, str2_comm), ut.compare(str1, str2_1c)])

                #print(coef_Lev, i_Lev)
                #print(coef_Jac, i_Jac)
                #print(coef_equ, i_equ)

                if coef_equ == 1:
                    if coef_Lev > 0.9 and coef_Jac > 0.6:
                        result = 2
                        name_in_base = str_target
                        break
                    elif coef_Lev > 0.6 or coef_Jac > 0.3:
                        result = 1
                        name_in_base = str_target
                        break
                    else:
                        components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
                        components_Jac.append([coef_Jac, str_target, strs[i_Jac]])
                        continue

                
                
                else:

                    if ut.isRLC(str1, options)[0]:
                        str1_offmask = ut.isRLC(str1, options)
                        match str1_offmask[1]:
                            case "R":
                                str2_offmask = ut.isRLC(str2_1c, options, "TYPE-SIZE-VALUE TOL")
                                if str2_offmask[0]  and str2_offmask[1] == "R":

                                    if ut.parse_component_value(str1_offmask[3].split(" ")[0]) != 0 and ut.parse_component_value(str2_offmask[3].split(" ")[0]) != 0:
                                        value_err = abs((ut.parse_component_value(str1_offmask[3].split(" ")[0]) / ut.parse_component_value(str2_offmask[3].split(" ")[0])) - 1)
                                    elif ut.parse_component_value(str1_offmask[3].split(" ")[0]) == ut.parse_component_value(str2_offmask[3].split(" ")[0]):
                                        value_err = 0
                                    else:
                                        value_err = 1


                                    if str1_offmask[1] == str2_offmask[1] and str1_offmask[2] == str2_offmask[2] and (value_err < 0.2):
                                        if coef_Lev > 0.6 or coef_Jac > 0.3:
                                            components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
                                            components_Jac.append([coef_Jac, str_target, strs[i_Jac]])
                                    else:
                                        continue
                            case "C":
                                str2_offmask = ut.isRLC(str2_1c, options, "TYPE-SIZE-DIE-VOL-VALUE TOL")
                                if str2_offmask[0] and str2_offmask[1] == "C":
                                    
                                    print(str1, str2_1c)
                                    print(str1_offmask, str2_offmask)
                                    value_err = abs((ut.parse_component_value(str1_offmask[5].split(" ")[0]) / ut.parse_component_value(str2_offmask[5].split(" ")[0])) - 1)
                                    print(value_err)
                                    

                                    if str1_offmask[1] == str2_offmask[1] and str1_offmask[2] == str2_offmask[2] and str1_offmask[4] <= str2_offmask[4] and (value_err < 0.2):
                                        if coef_Lev > 0.6 or coef_Jac > 0.3:
                                            components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
                                            components_Jac.append([coef_Jac, str_target, strs[i_Jac]])
                                    else:
                                        continue

                    else:
                        if coef_Lev > 0.5:
                            components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
                        elif coef_Jac > 0.5:
                            components_Jac.append([coef_Jac, str_target, strs[i_Jac]])


            #     if coef_equ == 1:
            #         if ut.compare(str1, str_target, 'Levenshtein') > 0.9 and ut.compare(str1, str_target, 'Jacquard') > 0.6:
            #             result = 2
            #             name_in_base = str_target
            #             break
            #         elif ut.compare(str1, str_target, 'Jacquard') > 0.3:
            #             result = 1
            #             name_in_base = str_target
            #             break
            #         else:
            #             components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
            #     if (coef_Lev > 0.5) and (coef_Jac > 0.5): 
            #         components_All.append([coef_Lev, coef_Jac, str_target, strs[i_Lev]])
            #     else:
            #         if (coef_Lev > 0.5):
            #             if ut.isRLC(str1)[0]:
            #                 if ut.isRLC(str1)[1] in str_target:
            #                     components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
            #             else:
            #                 components_Lev.append([coef_Lev, str_target, strs[i_Lev]])
            #         if (coef_Jac > 0.5):
            #             components_Jac.append([coef_Jac, str_target, strs[i_Jac]])

            # print(components_Lev)
            # print()

            components_Lev = sorted(components_Lev, key=lambda x: x[0], reverse=True)
            components_Jac = sorted(components_Jac, key=lambda x: x[0], reverse=True)
            # components_All = sorted(components_All, key=lambda x: x[0], reverse=True)

            # print(components_Lev)

            if result == 0:
                if options.compare_method == "normal":
                    if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                        if components_Lev[0][0] > 0.9 and components_Jac[0][0] > 0.6:
                            result = 1
                            name_in_base = str(components_Lev[0][1])
                        else:
                            if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                                if components_Lev[0][0] > 0.95:
                                    result = 1
                                    name_in_base = str(components_Lev[0][1])
                                else:
                                    result = -1
                            else:
                                result = -2
                    else:
                        result = -2
                    # if (len(components_All) > 0):
                    #     if components_All[0][0] > 0.9 and components_All[0][1] > 0.6:
                    #         result = 1
                    #         name_in_base = str(components_All[0][2])
                    #     else:
                    #         if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                    #             if components_Lev[0][0] > 0.95:
                    #                 result = 1
                    #                 name_in_base = str(components_Lev[0][1])
                    #             else:
                    #                 result = -1
                    #         else:
                    #             result = -2
                    # else:
                    #     if (len(components_Lev) > 0) or (len(components_Jac) > 0):
                    #         if components_Lev[0][0] > 0.95:
                    #             result = 1
                    #             name_in_base = str(components_Lev[0][1])
                    #         else:
                    #             result = -1
                    #     else:
                    #         result = -2

            if result < 0:
                bom_table.at[index_bom, 'SMT Левенштейна'] = str(components_Lev)
                bom_table.at[index_bom, 'SMT Жаккара'] = str(components_Jac)

            bom_table.at[index_bom, 'SMT'] = name_in_base
            bom_table.at[index_bom, 'Результат'] = result
            bom_table.at[index_bom, 'SMT Баланс'] = balance
        i += 1
        proc = i / (len(bom_table) + 1) * 100
        options.log_object.update_bar_signal.emit(proc)

        if not(options.quiet_mode):
            if str1 != "" and str1 != " ":
                if (len(components_Lev) > 0):
                    lev = components_Lev[0][0]
                else:
                    lev = -1
                st.log(f"target={str1}, result={name_in_base}, c_lev={lev}, status={str(result)}, balance={str(balance)}.", options.log_object)

    end_time = dt.datetime.now()
    deltatime_str = st.get_time_difference(start_time, end_time)
    st.log(f"The LIST file with results has been generated! Lead time: {deltatime_str}", options.log_object)
    if options.save_list_after_check:
        bom_table.to_excel(options.name_list_after_check, index=False)
        st.log(f"The LIST file with the results was saved as {options.name_list_after_check}.", options.log_object)


    return bom_table
        


def draw_file(name_bom, bom_table, outputname="output.xlsx", open_file=False, options=-1, preset_list='LIST_DEFAULT'):

    options = add_config_to_class(options, preset_list)

    st.log("The process draw file has begun.", options.log_object)
    st.log(f"Options: name_bom={name_bom}, output_name={outputname}, open_file={open_file}", options.log_object)
    start_time = dt.datetime.now()


    bom_table_ = pd.read_excel(name_bom, usecols=options.list_use_columns, skiprows=options.list_skip_rows)
    header_bom = bom_table_.columns.tolist()
    name_pn_bom = header_bom[options.list_pn_number]
    
    workbook = oxl.load_workbook(name_bom)

    main_sheet = workbook.worksheets[0]
    main_sheet.title = "Перечень элементов"

    color2_fill = PatternFill(start_color="97e253", end_color="97e253", fill_type="solid")
    color1_fill = PatternFill(start_color="ccf1aa", end_color="ccf1aa", fill_type="solid") 
    color0_fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    color_1_fill = PatternFill(start_color="ffd889", end_color="ffd889", fill_type="solid")
    color_2_fill = PatternFill(start_color="f4a190", end_color="f4a190", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),  # Левая граница
        right=Side(style="thin"),  # Правая граница
        top=Side(style="thin"),  # Верхняя граница
        bottom=Side(style="thin")  # Нижняя граница
    )

    if options.moves:
        ut.move_column(main_sheet, 12, 1)
        ut.move_column(main_sheet, 4, 50, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 5, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 6, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 7, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 8, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 9, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 10, -1, skip=options.list_skip_rows-1)
        ut.move_column(main_sheet, 54, -44, skip=options.list_skip_rows-1)

        main_sheet.cell(row=options.list_skip_rows+1, column=13).alignment = alignment_center

    main_sheet.cell(row=options.list_skip_rows+1, column=options.start_custom_column).value = "SMT"
    main_sheet.cell(row=options.list_skip_rows+1, column=options.start_custom_column).border = thin_border

    main_sheet.cell(row=options.list_skip_rows+1, column=options.start_custom_column+1).value = "SMT Баланс"
    main_sheet.cell(row=options.list_skip_rows+1, column=options.start_custom_column+1).border = thin_border

    if options.frame_visible:
        table_column = 9

        main_sheet.cell(row=3, column=table_column).fill = color2_fill
        main_sheet.cell(row=3, column=table_column+1).value = "Наименование есть в SMT"
        main_sheet.cell(row=3, column=table_column+1).alignment = alignment_left
        main_sheet.cell(row=4, column=table_column).fill = color1_fill
        main_sheet.cell(row=4, column=table_column+1).value = "Наименование есть в SMT, но написание отличается"
        main_sheet.cell(row=4, column=table_column+1).alignment = alignment_left
        main_sheet.cell(row=5, column=table_column).fill = color0_fill
        main_sheet.cell(row=5, column=table_column+1).value = "Не анализируется"
        main_sheet.cell(row=5, column=table_column+1).alignment = alignment_left
        main_sheet.cell(row=6, column=table_column).fill = color_1_fill
        main_sheet.cell(row=6, column=table_column+1).value = "Наименование не найдено в SMT, но есть похожие"
        main_sheet.cell(row=6, column=table_column+1).alignment = alignment_left
        main_sheet.cell(row=7, column=table_column).fill = color_2_fill
        main_sheet.cell(row=7, column=table_column+1).value = "Наименование не найдено в SMT"
        main_sheet.cell(row=7, column=table_column+1).alignment = alignment_left

        for i in range(3, 8):
            for j in range(table_column, 12):
                main_sheet.cell(row=i, column=j).border = thin_border

            main_sheet.merge_cells(f'J{i}:L{i}')

    for index_bom, row_bom in bom_table.iterrows():

        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column).value = bom_table.at[index_bom, 'SMT']
        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column).alignment = alignment_center
        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column).border = thin_border

        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column+1).value = bom_table.at[index_bom, 'SMT Баланс']
        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column+1).alignment = alignment_center
        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column+1).border = thin_border

        main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=13).alignment = alignment_center

        match bom_table.at[index_bom, 'Результат']:
            case -1:
                title_ = row_bom[name_pn_bom].replace("/", "")
                if len(title_) > 30:
                    title_ = title_[:29]
                new_sheet = workbook.create_sheet(title=title_)
                
                hyperlink_cell1 = new_sheet.cell(row=1, column=1)  
                hyperlink_cell1.value = "На главную"
                hyperlink_cell1.hyperlink = f"#'{main_sheet.title}'!{get_column_letter(options.start_custom_column)}{index_bom+options.list_skip_rows + 2}" 
                hyperlink_cell1.font = Font(underline="single", color="0563C1")  
                new_sheet.cell(row=1, column=1).alignment = alignment_center

                hyperlink_cell2 = main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=options.start_custom_column)  
                hyperlink_cell2.value = str(eval(row_bom['SMT Левенштейна'])[0][1]).replace("/", "").replace(".", "_") #"Похожие"
                hyperlink_cell2.hyperlink = f"#'{title_}'!A1" 
                hyperlink_cell2.font = Font(underline="single", color="0563C1") 
                row = 3
                new_sheet.cell(row=row, column=1).value = "Референс:"
                new_sheet.cell(row=row, column=1).font = Font(bold=True)
                new_sheet.cell(row=row, column=1).alignment = alignment_center
                new_sheet.cell(row=row, column=1).border = thin_border
                new_sheet.cell(row=row, column=2).value = str(row_bom[name_pn_bom])
                new_sheet.cell(row=row, column=2).alignment = alignment_center
                new_sheet.cell(row=row, column=2).border = thin_border
                new_sheet.row_dimensions[row].height = 25.5
                row = 5
                new_sheet.cell(row=row, column=1).value = "Коэффициент Левенштейна"
                new_sheet.cell(row=row, column=2).value = "SMT"
                new_sheet.cell(row=row, column=1).font = Font(bold=True)
                new_sheet.cell(row=row, column=2).font = Font(bold=True)
                new_sheet.cell(row=row, column=1).alignment = alignment_center
                new_sheet.cell(row=row, column=1).border = thin_border
                new_sheet.cell(row=row, column=2).alignment = alignment_center
                new_sheet.cell(row=row, column=2).border = thin_border
                new_sheet.row_dimensions[row].height = 25.5
                
                row += 1
                for elem in eval(row_bom['SMT Левенштейна']):
                    new_sheet.cell(row=row, column=1).value = elem[0]
                    new_sheet.cell(row=row, column=2).value = elem[1]
                    
                    new_sheet.cell(row=row, column=1).alignment = alignment_center
                    new_sheet.cell(row=row, column=1).border = thin_border
                    new_sheet.cell(row=row, column=2).alignment = alignment_center
                    new_sheet.cell(row=row, column=2).border = thin_border

                    row += 1

                ut.set_column_autowidth(new_sheet, ['A', 'B'])
            case -2:
                pass

        for i in range(1, options.start_custom_column+2):
            match bom_table.at[index_bom, 'Результат']:
                case 2:
                    main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=i).fill = color2_fill
                case 1:
                    main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=i).fill = color1_fill
                case 0:
                    main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=i).fill = color0_fill
                case -1:
                    main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=i).fill = color_1_fill
                case -2:
                    main_sheet.cell(row=index_bom + options.list_skip_rows + 2, column=i).fill = color_2_fill

    if options.widths:
        main_sheet.column_dimensions['F'].width = main_sheet.column_dimensions['G'].width
        main_sheet.column_dimensions['G'].width = main_sheet.column_dimensions['H'].width
        main_sheet.column_dimensions['H'].width = main_sheet.column_dimensions['I'].width
        main_sheet.column_dimensions['I'].width = main_sheet.column_dimensions['J'].width
        ut.set_column_autowidth(main_sheet, ['J', 'D', 'L'])
        main_sheet.column_dimensions['J'].width = main_sheet.column_dimensions['K'].width

    ut.set_column_autowidth(main_sheet, [f'{get_column_letter(options.start_custom_column+1)}'])
    ut.set_column_autowidth(main_sheet, [f'{get_column_letter(options.start_custom_column)}'], 1.1)
    ut.set_column_autowidth(main_sheet, [f'{get_column_letter(options.start_custom_column+2)}'], 1.1)

    outputname = st.get_name_file(outputname)

    workbook.save(outputname)
    if open_file:
        try:
            os.startfile(outputname)
        except:
            st.log(f"The file cannot be opened. It may be busy! File: {outputname}", options.log_object)


    end_time = dt.datetime.now()
    deltatime_str = st.get_time_difference(start_time, end_time)
    st.log(f"The draw file has been generated! Lead time: {deltatime_str}", options.log_object)
    options.log_object.update_bar_signal.emit(100)

def search(name_base, search_line, options, preset_base='BASE_DEFAULT'):

    print(name_base)
    print(search_line)
    
    start_time = dt.datetime.now()
    base_table = pd.read_html(name_base, encoding='cp1251')[int(options.config[preset_base]['base_table_number'])]

    header_base = base_table.columns.tolist()
    #for i_column in options.base_drop_column:
    #    base_table = base_table.drop(columns=[header_base[i_column]])
    base_table = base_table.fillna("")

    output_table = pd.DataFrame()
    i = 1
    j = 1
    last_proc = 0

    for index, row in base_table.iterrows():
        if not(options.log_object._is_running):
            break

        i += 1

        if fnmatch.fnmatch(row['Component'], f"*{search_line}*") or fnmatch.fnmatch(row['Comment'], f"*{search_line}*") or fnmatch.fnmatch(row[header_base[6]], f"*{search_line}*"):
            #output_table = output_table._append(row, ignore_index=True)
            output_table = pd.concat([output_table, row], axis=1, ignore_index=True)

            model = PandasModel(output_table)
            time.sleep(0.2)
            options.log_object.update_table_signal.emit(model)
            j += 1

        proc = round(i/(len(base_table)+1) * 10)

        if not(last_proc == proc):
            print(proc)
            options.log_object.update_bar_signal.emit(proc*10)
        
        last_proc = proc

        if j == 100:
            break

    options.log_object.update_bar_signal.emit(100)

if __name__ == "__main__":

    name_bom = ""
    name_base = ""
    options_1 = options()

    bom_table = find_bom_in_base(name_bom, name_base, options_1)
    draw_file(name_bom, bom_table)
